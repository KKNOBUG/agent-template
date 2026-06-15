# -*- coding: utf-8 -*-
"""
@Author  : weixianzhe
@Project : KeenRobot
@Module  : xlsx_writer.py
@DateTime: 2026/6/11
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLSX_HEADERS = [
    "测试用例ID", "测试用例标题", "测试用例描述", "正/反案例",
    "测试步骤", "前置条件", "预期结果", "优先级", "测试类别", "备注",
]


def parse_md(filepath: str) -> list[tuple[str, str, dict]]:
    """解析功能测试用例 Markdown 文件。

    返回 [(模块名, 测试点名, 用例字典), ...]。
    """
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    results: list[tuple[str, str, dict]] = []
    current_module = ""
    current_test_point = ""

    for line in lines:
        stripped = line.strip()

        m = re.match(r"^##\s+(模块\d+[：:]\s*.+)", stripped)
        if m:
            current_module = m.group(1).strip()
            continue

        m = re.match(r"^###\s+(测试点\d+[：:]\s*.+)", stripped)
        if m:
            current_test_point = m.group(1).strip()
            continue

        if not stripped.startswith("|"):
            continue
        if re.match(r"^\|[\s\-:]+\|", stripped):
            continue
        if all(h in stripped for h in ["测试用例ID", "测试用例标题"]):
            continue

        cells = [c.strip() for c in stripped.strip("|").split("|")]
        if len(cells) < len(XLSX_HEADERS):
            continue

        row = dict(zip(XLSX_HEADERS, cells[: len(XLSX_HEADERS)]))
        results.append((current_module, current_test_point, row))

    return results


def md_to_xlsx(input_path: str, output_path: str | None = None) -> str:
    """将功能测试用例 Markdown 文件直接转换为格式化的 XLSX 文件。

    自动识别 ## 模块 和 ### 测试点 层级，按模块/测试点分组展示，
    正案例绿色底、反案例红色底，带自适应行高。

    Args:
        input_path: Markdown 文件路径。
        output_path: 输出 XLSX 路径，默认为同目录同名 .xlsx。

    Returns:
        生成的 XLSX 文件绝对路径。
    """
    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".xlsx"

    rows = parse_md(input_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "功能测试用例"

    module_font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    point_font = Font(name="微软雅黑", size=11, bold=True, color="2E75B6")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(wrap_text=True, vertical="top", horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    positive_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    negative_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    md_col_widths = [16, 40, 40, 10, 60, 50, 60, 8, 10, 30]
    for idx, w in enumerate(md_col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    row_num = 1
    current_module = ""
    current_point = ""

    for mod, point, data in rows:
        if mod != current_module:
            current_module = mod
            current_point = ""
            ws.cell(row=row_num, column=1, value=mod).font = module_font
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(XLSX_HEADERS))
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center")
            row_num += 1

        if point != current_point:
            current_point = point
            ws.cell(row=row_num, column=1, value=point).font = point_font
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(XLSX_HEADERS))
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center")
            row_num += 1
            for col_idx, h in enumerate(XLSX_HEADERS, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1

        is_positive = data["正/反案例"] == "正案例"
        for col_idx, h in enumerate(XLSX_HEADERS, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=data[h])
            cell.font = cell_font
            cell.alignment = wrap_align if h not in ("测试用例ID", "正/反案例", "优先级", "测试类别") else center_align
            cell.border = thin_border
            cell.fill = positive_fill if is_positive else negative_fill

        max_lines = 1
        for h in ("测试步骤", "前置条件", "预期结果"):
            lines = data[h].count("<br>") + 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row_num].height = max(30, max_lines * 16)
        row_num += 1

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return os.path.abspath(output_path)
